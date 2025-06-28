import os
import asyncio
from datetime import datetime, UTC

import pytz
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from sqlalchemy.ext.asyncio import AsyncSession

from openpyxl import Workbook
from openpyxl.styles import Font

from bot import bot, keyboards
from bot.states import Mailing
from services.filters import IsAdmin
from database.repository.user import UserRepository

router = Router()


@router.callback_query(F.data == 'cancel')
async def cancel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()


@router.message(Command('admin'), IsAdmin())
async def admin_commands(message: Message):
    await message.answer(
        '/ban - <i>Бан пользователя</i>\n'
        '/unban - <i>Разбан пользователя</i>\n'
        '/statistics - <i>Статистика</i>\n'
        '/mailing - <i>Рассылка пользователям</i>'
    )


@router.message(Command('ban'), IsAdmin())
async def ban_user(message: Message, session: AsyncSession):
    splited_text = message.text.split()

    if len(splited_text) == 1:
        await message.answer(
            '<i>Для бана пользователя введите команду в таком формате</i>\n'
            '\n'
            '/ban [Telegram ID пользователя]'
        )
    elif len(splited_text) == 2:
        _, user_id = splited_text
        if user_id.isdigit():
            user_repo = UserRepository(session)
            user_id = int(user_id)

            if await user_repo.is_exist(user_id):
                if await user_repo.is_blocked(user_id):
                    await message.answer('Данный пользователь был забанен ранее')
                else:
                    await user_repo.ban(user_id)
                    await message.answer('Пользователь успешно забанен!')
            else:
                await message.answer('Данного пользователя нет в Базе')
        else:
            await message.answer('Введите <b><u>Telegram ID</u></b> пользователя которого хотите забанить')
    else:
        await message.answer(
            '<i>Для бана пользователя введите команду в таком формате</i>\n'
            '\n'
            '/ban [Telegram ID пользователя]'
        )


@router.message(Command('unban'), IsAdmin())
async def unban_user(message: Message, session: AsyncSession):
    splited_text = message.text.split()

    if len(splited_text) == 1:
        await message.answer(
            '<i>Для разбана пользователя введите команду в таком формате</i>\n'
            '\n'
            '/unban [Telegram ID пользователя]'
        )
    elif len(splited_text) == 2:
        _, user_id = splited_text
        if user_id.isdigit():
            user_repo = UserRepository(session)
            user_id = int(user_id)

            if await user_repo.is_exist(user_id):
                if await user_repo.is_blocked(user_id):
                    await user_repo.unban(user_id)
                    await message.answer('Пользователь был успешно разблокирован!')
                    await bot.send_message(
                        user_id,
                        '<b>Вы были разбанены!</b>'
                    )
                else:
                    await message.answer('Пользователь не в бане')
            else:
                await message.answer('Данного пользователя нет в Базе')
        else:
            await message.answer('Введите <b><u>Telegram ID</u></b> пользователя которого хотите разбанить')
    else:
        await message.answer(
            '<i>Для разбана пользователя введите команду в таком формате</i>\n'
            '\n'
            '/unban [Telegram ID пользователя]'
        )


@router.message(Command('statistics'), IsAdmin())
async def statistics(message: Message, session: AsyncSession):
    user_repo = UserRepository(session)

    stats = await user_repo.get_statistics()

    await message.answer(
        f'📊 <b>Статистика запусков бота</b>\n'
        f'\n'
        f'🔹 <i>За сегодня</i>: <b>{stats['today']}</b>\n'
        f'🔹 <i>За вчера</i>: <b>{stats['yesterday']}</b>\n'
        f'🔹 <i>За 3 дня</i>: <b>{stats['last_3_days']}</b>\n'
        f'🔹 <i>За 7 дней</i>: <b>{stats['last_7_days']}</b>',
        reply_markup=keyboards.uploading_users()
    )


@router.callback_query(F.data == 'uploading_users', IsAdmin())
async def uploading_users(callback: CallbackQuery, session: AsyncSession):
    msg = await callback.message.answer('<b>Готовим файл...</b>')

    try:
        user_repo = UserRepository(session)
        users = await user_repo.get_all_users()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Список пользователей'

        headers = ['ID', 'username', 'Имя', 'Фамилия', 'Дата запуска (МСК)', 'Заблокирован']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user.id)
            ws.cell(row=row_num, column=2, value=user.username)
            ws.cell(row=row_num, column=3, value=user.first_name)
            ws.cell(row=row_num, column=4, value=user.last_name)
            ws.cell(row=row_num, column=5, value=user.started_at.astimezone(pytz.timezone('Europe/Moscow')).strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row_num, column=6, value='Да' if user.is_blocked else 'Нет')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        filename = f'users_export_{datetime.now(UTC).astimezone(pytz.timezone('Europe/Moscow')).strftime('%d-%m-%Y_%H-%M')}.xlsx'
        wb.save(filename)

        await callback.message.answer_document(
            document=FSInputFile(path=filename)
        )

        if os.path.exists(filename):
            os.remove(filename)

    except Exception:
        await callback.message.answer('<b>Ошибка при формирования таблицы</b>')
    finally:
        await msg.delete()


@router.message(Command('mailing'), IsAdmin())
async def mailing(message: Message, state: FSMContext):
    await message.answer(
        'Введите сообщение которое хотели бы разослать всем пользователям...',
        reply_markup=keyboards.cancel()
    )
    await state.set_state(Mailing.message)


@router.message(F.text | F.photo | F.video, IsAdmin(), Mailing.message)
async def mailing_warning(message: Message, state: FSMContext):
    msg = await bot.copy_message(
        chat_id=message.from_user.id,
        from_chat_id=message.from_user.id,
        message_id=message.message_id,
    )
    await bot.send_message(
        chat_id=message.from_user.id,
        text=(
            'Вы уверены что хотите отправить данное сообщение всем пользователям?\n'
            '\n'
            '<i>Если хотите изменить сообщение, просто отправьте новое сообщение</i>'
         ),
        reply_markup=keyboards.mailing_confirm(),
        reply_to_message_id=msg.message_id
    )
    await state.update_data(mail_msg_id=message.message_id)


@router.callback_query(F.data == 'mailing_confirm', Mailing.message, IsAdmin())
async def mailing_run(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    await callback.message.delete()

    data = await state.get_data()
    await state.clear()

    user_repo = UserRepository(session)

    info_msg = await callback.message.answer('<b>Рассылка началась...</b>')
    orig_msg_id: int = data['mail_msg_id']

    user_ids = await user_repo.get_all_users(only_id=True)

    success = 0
    un_success = 0
    for i, user_id in enumerate(user_ids, start=1):
        try:
            await bot.copy_message(
                chat_id=user_id,
                from_chat_id=callback.from_user.id,
                message_id=orig_msg_id
            )
            success += 1
        except Exception:
            un_success += 1
            pass
        
        if i % 15 == 0:
            await asyncio.sleep(3)
            await info_msg.edit_text(
                f'<b>Рассылка началась...</b> ({i}/{len(user_ids)})\n'
                f'\n'
                f'Успешно: {success}\n'
                f'Не успешно: {un_success}'
            )

    await info_msg.delete()
    await callback.message.answer(
        f'<b>Рассылка завершена</b>\n'
        f'\n'
        f'Успешно: {success}\n'
        f'Не успешно: {un_success}'
    )
